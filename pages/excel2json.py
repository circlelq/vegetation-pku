# !/usr/bin/python
#  -*- coding: UTF-8 -*-
import json
from PIL.Image import new
import openpyxl
import argparse
import sys
import os
import time
import urllib.parse
import urllib

today = time.strftime("%Y-%m-%d", time.localtime())


workbook = openpyxl.load_workbook('植物档案.xlsx')

data = workbook.active

rowNum = data.max_row  # sheet行数
colNum = data.max_column   # sheet列数

#  获取所有单元格的内容
data_list = []
for i in range(1, rowNum):
    rowlist = []
    for j in range(colNum):
        cell = data.cell(i+1, j+1).value
        if cell == None:
            cell = ''
        rowlist.append(cell)
    data_list.append(rowlist)


# 输出所有单元格的内容

rowNum -= 1


# for i in range(rowNum):
#     for j in range(colNum):
#         print(data_list[i][j], ' ', end="")
#     print('\n')


labels = [
    [0, 'name', lambda x:x],
    [1, 'latin', lambda x:x],
    [2, 'photos', lambda x:x],
    [3, 'Genus', lambda x:x],
    [4, 'Family', lambda x:x],
    [5, 'Order', lambda x:x],
    [6, 'Class', lambda x:x],
    [7, 'Phylum', lambda x:x],
    [8, 'viewingTime', lambda x:x],
    [9, 'places', lambda x:x],
    [10, 'markers', lambda x:x],
    [11, 'relation', lambda x:x],


    # [9, '状况', lambda x:'不明' if len(x) < 1 else x],

]

data_json = []

for i in range(rowNum):
    json_line = {}

    for j in labels:
        json_line[j[1]] = j[2](data_list[i][j[0]])

    data_json.append(json_line)


# 编写 data base

with open('db.json', 'w', encoding='utf-8') as f:
    for line in data_json:
        if line['photos'] == '':
            continue
        mapid = 0
        markers = {}
        maps = line['markers']
        coor = maps.split(";")
        coor.remove("")
        for point in coor:
            tempPoint = {}
            x, y = point.split(",")
            tempPoint["coordinates"] = [y, x]
            tempPoint['type'] = 'Point'
            markers[str(mapid)] = tempPoint
            mapid = mapid + 1
        line['markers'] = markers
        newLine = line.copy()
        for i in line:
            if line[i] == "" or line[i] == {}:
                del newLine[i]
        f.write(json.dumps(newLine, ensure_ascii=False))
